# rent/views/report_t_views.py


"""
✅ Updated to use ContractFinancialService
تقارير المستأجرين - محدث لاستخدام الخدمة الموحدة
"""

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required, permission_required
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ✅ NEW: استيراد الخدمة الموحدة
from rent.models.contract_models import Contract
from rent.services.contract_financial_service import (
    ContractFinancialService,
    generate_tenants_report,  # ✅ دالة التوافق من الخدمة الموحدة
)


# ========================================
# دوال مساعدة
# ========================================

def get_contracts_queryset(filters=None):
    """
    الحصول على العقود مع الفلاتر
    يشمل: النشطة + المنتهية + الملغاة (لظهور المستحقات المتبقية)
    """
    queryset = Contract.objects.filter(
        status__in=['active', 'expired', 'terminated']
    ).select_related(
        'tenant',
    ).prefetch_related(
        'units',
        'modifications',
        'receipts'
    )
    
    # تطبيق الفلاتر الإضافية
    if filters:
        if filters.get('tenant_name'):
            queryset = queryset.filter(
                tenant__name__icontains=filters['tenant_name']
            )
        
        if filters.get('location'):
            queryset = queryset.filter(
                Q(units__location__icontains=filters['location']) |
                Q(units__building__location__icontains=filters['location'])
            ).distinct()
    
    return queryset


def apply_outstanding_filter(report_data, has_outstanding_filter):
    """
    فلترة حسب وجود مستحقات
    """
    if not has_outstanding_filter:
        return report_data
    
    if has_outstanding_filter == 'yes':
        return [r for r in report_data if r['outstanding_amount'] > 0]
    elif has_outstanding_filter == 'no':
        return [r for r in report_data if r['outstanding_amount'] == 0]
    
    return report_data


# ========================================
# View 1: التقرير الأساسي
# ========================================

@login_required
@permission_required('rent.view_reports', raise_exception=True)
def tenants_report_view(request):
    """
    عرض تقرير المستأجرين الأساسي
    ✅ Updated - يستخدم ContractFinancialService
    """
    # جلب العقود (نشطة + منتهية + ملغاة) لظهور المستحقات المتبقية
    contracts = Contract.objects.filter(
        status__in=['active', 'expired', 'terminated']
    ).select_related(
        'tenant',
    ).prefetch_related(
        'units',
        'modifications',
        'receipts'
    )
    
    # استخدام الخدمة الموحدة
    report_data = generate_tenants_report(contracts)

    # عرض العقود التي عليها استحقاق مالي فقط (مستحق السداد)
    report_data = [r for r in report_data if r['outstanding_amount'] > 0]
    # ترتيب: حسب تاريخ الاستحقاق (الأقدم أولاً)
    report_data.sort(key=lambda x: x.get('due_period_from') or date.min)

    # حساب الإجماليات
    total_outstanding = sum(
        item['outstanding_amount'] for item in report_data
    )

    total_annual_rent = sum(
        item['annual_rent'] for item in report_data
    )

    # إحصائيات إضافية
    contracts_with_outstanding = len(report_data)

    context = {
        'report_data': report_data,
        'total_outstanding': total_outstanding,
        'total_annual_rent': total_annual_rent,
        'contracts_count': len(report_data),
        'contracts_with_outstanding': contracts_with_outstanding,
        'report_date': date.today(),
    }
    
    return render(request, 'reports/tenants_report.html', context)


# ========================================
# View 2: التقرير المفصّل مع الفلاتر
# ========================================

@login_required
@permission_required('rent.view_reports', raise_exception=True)
def detailed_tenants_report(request):
    """
    تقرير مفصّل مع إمكانية الفلترة
    ✅ Updated - يستخدم ContractFinancialService
    """
    # جمع الفلاتر من الطلب
    filters = {
        'tenant_name': request.GET.get('tenant_name', ''),
        'location': request.GET.get('location', ''),
        'has_outstanding': request.GET.get('has_outstanding', ''),
    }
    
    # جلب العقود مع الفلاتر
    contracts = get_contracts_queryset(filters)
    
    # ✅ NEW: استخدام الخدمة الموحدة
    report_data = generate_tenants_report(contracts)
    
    # تطبيق فلتر المستحقات (فلتر العرض)
    report_data = apply_outstanding_filter(
        report_data, 
        filters.get('has_outstanding')
    )
    
    # حساب الإحصائيات
    stats = {
        'total_contracts': len(report_data),
        'total_outstanding': sum(item['outstanding_amount'] for item in report_data),
        'total_annual_rent': sum(item['annual_rent'] for item in report_data),
        'contracts_with_outstanding': sum(
            1 for item in report_data if item['outstanding_amount'] > 0
        ),
        'total_overdue_periods': sum(
            item['overdue_periods_count'] for item in report_data
        ),
    }
    
    context = {
        'report_data': report_data,
        'stats': stats,
        'filters': filters,
        'report_date': date.today(),
    }
    
    return render(request, 'reports/detailed_tenants_report.html', context)


# ========================================
# View 3: تصدير إلى Excel
# ========================================

@login_required
@permission_required('rent.export_reports', raise_exception=True)
def export_tenants_report_excel(request):
    """
    تصدير التقرير إلى Excel مع تنسيق احترافي
    ✅ Updated - يستخدم ContractFinancialService
    """
    # جلب الفلاتر (إن وجدت)
    filters = {
        'tenant_name': request.GET.get('tenant_name', ''),
        'location': request.GET.get('location', ''),
        'has_outstanding': request.GET.get('has_outstanding', ''),
    }
    
    # جلب العقود
    contracts = get_contracts_queryset(filters)
    
    # ✅ NEW: استخدام الخدمة الموحدة
    report_data = generate_tenants_report(contracts)
    
    # تطبيق فلتر المستحقات
    report_data = apply_outstanding_filter(
        report_data, 
        filters.get('has_outstanding')
    )
    
    # إنشاء Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير المستأجرين"
    
    # ========================================
    # تنسيق العناوين
    # ========================================
    
    headers = [
        'رقم المستأجر',
        'اسم المستأجر',
        'الموقع',
        'الإيجار السنوي',
        'رقم المعرض',
        'الهاتف',
        'المستحق',
        'من تاريخ',
        'إلى تاريخ',
        'عدد الفترات المتأخرة'
    ]
    
    # تنسيق الحدود
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # كتابة العناوين
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        cell.border = thin_border
    
    # ========================================
    # كتابة البيانات
    # ========================================
    
    row_num = 2
    total_outstanding = Decimal('0')
    total_annual_rent = Decimal('0')
    
    for item in report_data:
        # رقم المستأجر
        ws.cell(row=row_num, column=1).value = item['tenant_id']
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=1).border = thin_border

        # اسم المستأجر
        ws.cell(row=row_num, column=2).value = item['tenant_name']
        ws.cell(row=row_num, column=2).border = thin_border

        # الموقع
        ws.cell(row=row_num, column=3).value = item['location']
        ws.cell(row=row_num, column=3).border = thin_border

        # الإيجار السنوي
        ws.cell(row=row_num, column=4).value = float(item['annual_rent'])
        ws.cell(row=row_num, column=4).number_format = '#,##0.00'
        ws.cell(row=row_num, column=4).border = thin_border

        # رقم المعرض (كل الوحدات)
        ws.cell(row=row_num, column=5).value = item.get('all_unit_numbers_str', item['unit_number'])
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=5).border = thin_border

        # الهاتف
        ws.cell(row=row_num, column=6).value = item.get('tenant_phone', 'غير محدد')
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=6).border = thin_border

        # المستحق
        ws.cell(row=row_num, column=7).value = float(item['outstanding_amount'])
        ws.cell(row=row_num, column=7).number_format = '#,##0.00'
        ws.cell(row=row_num, column=7).border = thin_border

        # تلوين المستحق إذا كان > 0
        if item['outstanding_amount'] > 0:
            ws.cell(row=row_num, column=7).font = Font(color="FF0000", bold=True)

        # من تاريخ
        due_from = item.get('due_period_from')
        ws.cell(row=row_num, column=8).value = due_from.strftime('%Y-%m-%d') if due_from else '-'
        ws.cell(row=row_num, column=8).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=8).border = thin_border

        # إلى تاريخ
        due_to = item.get('due_period_to')
        ws.cell(row=row_num, column=9).value = due_to.strftime('%Y-%m-%d') if due_to else '-'
        ws.cell(row=row_num, column=9).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=9).border = thin_border

        # عدد الفترات المتأخرة
        ws.cell(row=row_num, column=10).value = item['overdue_periods_count']
        ws.cell(row=row_num, column=10).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=10).border = thin_border

        # الإجماليات
        total_outstanding += item['outstanding_amount']
        total_annual_rent += item['annual_rent']

        row_num += 1
    
    # ========================================
    # إضافة صف الإجماليات
    # ========================================
    
    ws.cell(row=row_num, column=3).value = "الإجمالي:"
    ws.cell(row=row_num, column=3).font = Font(bold=True, size=12)
    ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=row_num, column=3).fill = PatternFill(
        start_color="E7E6E6",
        end_color="E7E6E6",
        fill_type="solid"
    )

    # إجمالي الإيجار السنوي
    ws.cell(row=row_num, column=4).value = float(total_annual_rent)
    ws.cell(row=row_num, column=4).number_format = '#,##0.00'
    ws.cell(row=row_num, column=4).font = Font(bold=True, size=12)
    ws.cell(row=row_num, column=4).fill = PatternFill(
        start_color="E7E6E6",
        end_color="E7E6E6",
        fill_type="solid"
    )

    # خلايا فارغة
    for col in [5, 6]:
        ws.cell(row=row_num, column=col).fill = PatternFill(
            start_color="E7E6E6",
            end_color="E7E6E6",
            fill_type="solid"
        )

    # إجمالي المستحق
    ws.cell(row=row_num, column=7).value = float(total_outstanding)
    ws.cell(row=row_num, column=7).number_format = '#,##0.00'
    ws.cell(row=row_num, column=7).font = Font(bold=True, size=12, color="FF0000")
    ws.cell(row=row_num, column=7).fill = PatternFill(
        start_color="E7E6E6",
        end_color="E7E6E6",
        fill_type="solid"
    )

    # خلايا فارغة
    for col in [8, 9, 10]:
        ws.cell(row=row_num, column=col).fill = PatternFill(
            start_color="E7E6E6",
            end_color="E7E6E6",
            fill_type="solid"
        )
    
    # ========================================
    # ضبط عرض الأعمدة
    # ========================================
    
    column_widths = {
        1: 15,  # رقم المستأجر
        2: 30,  # اسم المستأجر
        3: 25,  # الموقع
        4: 18,  # الإيجار السنوي
        5: 20,  # رقم المعرض
        6: 18,  # الهاتف
        7: 18,  # المستحق
        8: 15,  # من تاريخ
        9: 15,  # إلى تاريخ
        10: 20, # عدد الفترات المتأخرة
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # تجميد الصف الأول
    ws.freeze_panes = 'A2'
    
    # ========================================
    # حفظ الملف
    # ========================================
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=tenants_report_{date.today().strftime("%Y%m%d")}.xlsx'
    )
    
    wb.save(response)
    
    return response


# ========================================
# View 4: تصدير إلى PDF
# ========================================

@login_required
@permission_required('rent.export_reports', raise_exception=True)
def export_tenants_report_pdf(request):
    """
    تصدير التقرير إلى PDF
    ✅ يستخدم xhtml2pdf
    """
    from django.template.loader import get_template
    from django.http import HttpResponse
    from io import BytesIO

    try:
        from xhtml2pdf import pisa
        HAS_XHTML2PDF = True
    except ImportError:
        HAS_XHTML2PDF = False

    # جلب الفلاتر
    filters = {
        'tenant_name': request.GET.get('tenant_name', ''),
        'location': request.GET.get('location', ''),
        'has_outstanding': request.GET.get('has_outstanding', ''),
    }

    # جلب العقود
    contracts = get_contracts_queryset(filters)

    # استخدام الخدمة الموحدة
    report_data = generate_tenants_report(contracts)

    # تطبيق فلتر المستحقات
    report_data = apply_outstanding_filter(
        report_data,
        filters.get('has_outstanding')
    )

    # فقط العقود التي عليها مستحقات
    report_data = [r for r in report_data if r['outstanding_amount'] > 0]

    # الإحصائيات
    stats = {
        'total_contracts': len(report_data),
        'total_outstanding': sum(item['outstanding_amount'] for item in report_data),
        'total_annual_rent': sum(item['annual_rent'] for item in report_data),
    }

    context = {
        'report_data': report_data,
        'stats': stats,
        'report_date': date.today(),
    }

    if HAS_XHTML2PDF:
        # استخدام xhtml2pdf لإنشاء PDF
        template = get_template('reports/tenants_report_pdf.html')
        html_string = template.render(context)

        # إنشاء PDF
        result = BytesIO()
        pdf = pisa.pisaDocument(
            BytesIO(html_string.encode('utf-8')),
            result,
            encoding='utf-8'
        )

        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = (
                f'attachment; filename=tenants_report_{date.today().strftime("%Y%m%d")}.pdf'
            )
            return response

    # Fallback: إرجاع HTML للطباعة
    return render(request, 'reports/tenants_report_pdf.html', context)


# ========================================
# View 5: طباعة التقرير (HTML للطباعة)
# ========================================

@login_required
@permission_required('rent.view_reports', raise_exception=True)
def print_tenants_report(request):
    """
    طباعة التقرير
    ✅ Updated - يستخدم ContractFinancialService
    """
    # جلب الفلاتر
    filters = {
        'tenant_name': request.GET.get('tenant_name', ''),
        'location': request.GET.get('location', ''),
        'has_outstanding': request.GET.get('has_outstanding', ''),
    }
    
    # جلب العقود
    contracts = get_contracts_queryset(filters)
    
    # ✅ NEW: استخدام الخدمة الموحدة
    report_data = generate_tenants_report(contracts)
    
    # تطبيق فلتر المستحقات
    report_data = apply_outstanding_filter(
        report_data,
        filters.get('has_outstanding')
    )
    
    # الإحصائيات
    stats = {
        'total_contracts': len(report_data),
        'total_outstanding': sum(item['outstanding_amount'] for item in report_data),
        'total_annual_rent': sum(item['annual_rent'] for item in report_data),
    }
    
    context = {
        'report_data': report_data,
        'stats': stats,
        'report_date': date.today(),
    }
    
    return render(request, 'reports/tenants_report_print.html', context)